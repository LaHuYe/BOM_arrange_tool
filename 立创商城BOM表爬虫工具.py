# coding=utf-8
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Side, Border, PatternFill, Alignment
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
import re

import time

from msedge.selenium_tools import EdgeOptions
from msedge.selenium_tools import Edge

import io
import requests
import zipfile
import winreg
# 立创商城URL
url = 'https://so.szlcsc.com/global.html?k='
# driver_path = r'D:\edgedriver\msedgedriver.exe'
driver_path = r'D:\edgedriver\msedgedriver.exe'
# 描述XPath
detailsLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[1]/div/ul/li[4]/div'
# 品牌XPath
brandLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[1]/div/ul/li[2]/a'
# 型号XPath
modelLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[1]/div/ul/li[1]/span[2]'
# 封装XPath
packageLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[1]/div/ul/li[3]/span[2]'
# 数量XPath
quantityLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[3]/div[1]/div[1]/ul/li[2]/div/p'
# # 价格XPath
# priceLabel = '//tbody/tr[1]/td/div[2]/div[2]/div[3]/div[1]/div[1]/ul/li[2]/div/span'
# 旧表头
old_label = [
    'Comment', 'Description', 'Designator', 'Footprint', 'LibRef', 'Quantity'  
]
# 新表头
label = [
    'Comment', 'Brand', 'Mdel', 'Details','Description','LibRef','Designator', 'Footprint', 'Number', 'Quantity', 'unitPrice', 'Total', 'Source'  # 填写表头
]
# 行宽度
ColumnWidth = [
    '15', '20', '25', '40','20','20','30', '25', '15', '10', '10', '10', '10'
]
# excel_name = "E:\Develop\Python\Python爬虫\python测试\ML302MOD_BOM_221212.xlsx"
excel_name = input("请输入文件绝对路径(例C:\\Users\\Wcy\\Desktop\\XXX)：")
if '"' in excel_name:
    excel_name = excel_name.replace('"', '')

if ".xlsx" in excel_name:
    wb = load_workbook(excel_name)
    excel_name = excel_name.replace(".xlsx", '')
else:
    wb = load_workbook(excel_name+str('.xlsx'))

sheets = wb.worksheets   # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]

def set_excel():
    newLabelFlag = False
    side = Side(style='thin', color='FF000000')  # 设置边框颜色
    border = Border(left=side, right=side, top=side, bottom=side)  # 设置边框样式

    for column in range(0, len(old_label)):
        if sheet1.cell(column=column+1, row=1).value == old_label[column]:
            oldLabelFlag = True
        else:
            oldLabelFlag = False
    if oldLabelFlag == True:
        # 删除列
        # sheet1.delete_cols(2)  # 删除第二列"Description"
        # sheet1.delete_cols(4)  # 删除第五列"Description"
        # 向下移动两行，向左移动一列
        # sheet1.move_range("B1:B"+str(sheet1.max_row), cols=1)
        # sheet1.move_range("E1:E"+str(sheet1.max_row), cols=-1)
        # 插入列
        sheet1.insert_cols(idx=2, amount=3)  # 从第2列插入，插入3列
        # sheet1.insert_cols(idx=8)  # 从第8列插入1列
        sheet1.move_range("H1:H"+str(sheet1.max_row), cols=2)#先移走
        sheet1.move_range("G1:G"+str(sheet1.max_row), cols=1)
        sheet1.move_range("F1:F"+str(sheet1.max_row), cols=1)
       
        sheet1.move_range("J1:J"+str(sheet1.max_row), cols=-4)#移回来
        sheet1.move_range("I1:I"+str(sheet1.max_row), cols=1)#移回来
        # sheet1.move_range("E1:E"+str(sheet1.max_row), cols=-1)
        # 填入表头数据
        for row in range(0, len(label)):
            sheet1.cell(column=row+1, row=1).value = label[row]
    else:
        for column in range(0, len(label)):
            if sheet1.cell(column=column+1, row=1).value == label[column]:
                newLabelFlag = True
            else:
                newLabelFlag = False
    if oldLabelFlag == True or newLabelFlag == True:
        data_all = sheet1.iter_cols(
            min_col=7, max_col=7, min_row=2, max_row=sheet1.max_row)
        for col in data_all:
            for cell in col:
                if sheet1.cell(row=cell.row, column=1).value == 'NC' \
                        or sheet1.cell(row=cell.row, column=1).value == 'TP' \
                        or sheet1.cell(row=cell.row, column=1).value == 'Test Point':
                    sheet1.delete_rows(cell.row)
       # 为每一列的每个单元格加边框
        for column in sheet1['A:M']:  # 设置A-K列加边框
            for cell in column:
                cell.border = border

        # 设置第一行填充颜色
        for row in sheet1[1]:
            row.fill = PatternFill(start_color="ffff00", fill_type="solid")

        # 字体居中并自动换行
        alignment_center = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        for area_column in range(1, sheet1.max_column + 1):
            for area_row in range(1, sheet1.max_row + 1):
                sheet1.cell(column=area_column,
                            row=area_row).alignment = alignment_center

        # 设置列宽
        for column in range(97, len(label)+97):  # 设置A-I列加边框
            sheet1.column_dimensions[chr(
                column)].width = ColumnWidth[column-97]

    else:
        print("EXCEL格式不对")


def open_browser():
    edge_options = EdgeOptions()
    edge_options.use_chromium = True
    edge_options.add_argument("--headless") #隐藏浏览器界面
    edge_options.add_argument('log-level=3')  #修改日等级，防止一直打印日志INFO:CONSOLE
    edge_options.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片, 提升速度
    edge_options.add_argument('--disable-gpu')  # 谷歌文档提到需要加上这个属性来规避bug
    # if getattr(sys, 'frozen', False):
    #     # 从exe包里找chromedriver依赖驱动的情况
    #     chromedriver_path = os.path.join(sys._MEIPASS, "msedgedriver.exe")
    #     driver = Edge(chromedriver_path,options=edge_options)
    # else:
    #     # 普通情况下从本地文件路径找依赖的情况
    #     driver = Edge(executable_path=r'D:\edgedriver\msedgedriver.exe',options=edge_options)
    return Edge(executable_path=driver_path,options=edge_options)

def download_msedgedriver():
    # 获取当前系统中安装的 Edge 浏览器版本号
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Edge\BLBeacon')
    version = winreg.QueryValueEx(key, 'version')[0]
    winreg.CloseKey(key)
    print(version)
    # 构造 msedgedriver 下载地址
    download_url = f'https://msedgedriver.azureedge.net/{version}/edgedriver_win64.zip'
    # 发送 GET 请求下载驱动程序
    response = requests.get(download_url)
    # 保存驱动程序到指定路径
    path = driver_path  # 替换为你想要保存的路径
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(zipfile.ZipFile(io.BytesIO(response.content)).read("msedgedriver.exe"))
    print(f"msedgedriver.exe 下载完成，保存路径为：{path}")


def get_data(): 
    empty_Flag = False
    data_all = sheet1.iter_cols(
        min_col=9, max_col=9, min_row=2, max_row=sheet1.max_row)
    for col in data_all:
        for cell in col:
            if (sheet1.cell(row=cell.row, column=9).value) != None:
                empty_Flag = True
        if empty_Flag == True:
            try:
                driver = open_browser()
            except WebDriverException as e:
                print(f"{driver_path} 不存在，正在下载...")
                download_msedgedriver()
                driver = open_browser()
            for cell in col:
                if (sheet1.cell(row=cell.row, column=9).value) == None:
                    for row in sheet1[cell.row]:
                        row.fill = PatternFill(
                            start_color="CCCCCC", fill_type="solid")
                    continue
                dj = 2
                driver.get(str(url)+cell.value)
                time.sleep(1)
                details = driver.find_elements_by_xpath(detailsLabel)
                brand = driver.find_elements_by_xpath(brandLabel)
                mdel = driver.find_elements_by_xpath(modelLabel)
                package = driver.find_elements_by_xpath(packageLabel)
                quantity = driver.find_elements_by_xpath(
                    quantityLabel[:58]+str(dj)+quantityLabel[59:])
                try:
                    # 计算器件位号数量
                    site = sheet1.cell(row=cell.row, column=7)
                    DesignatorNum = site.value.count(
                            ',')+site.value.count('，')+1
                    # 根据器件数量计算价位，按1000套计算      
                    while int(re.sub(r'\s|\D', '', quantity[0].text)) < DesignatorNum * 1000:
                        dj += 1
                        position = quantityLabel[:58] + \
                            str(dj)+quantityLabel[59:]
                        quantity = driver.find_elements_by_xpath(position)
                        if quantity == []:
                            break
                    dj -= 1
                    position = quantityLabel[:58] + str(dj)+quantityLabel[59:]      
                    quantity = driver.find_elements_by_xpath(position[:65]+'span')


                    print('\n')
                    print(re.sub(r'\s', '', brand[0].text))
                    print(re.sub(r'\s', '', mdel[0].text))
                    print(re.sub(r'\s', '', package[0].text))
                    print(re.sub(r'\s|描述：', '', details[0].text))
                    print('单价：' + re.sub(r'\s|￥', '',quantity[0].text))
                    if re.sub(r'\s|￥', '', quantity[0].text) != '报价':
                        print('总价:', float(re.sub(r'\s|￥', '', quantity[0].text))*DesignatorNum)
                    else:
                        print('总价:', re.sub(r'\s|￥', '', quantity[0].text))
                    
                    sheet1.cell(row=cell.row, column=4).value = re.sub(#描述
                            r'\s|描述：', '', details[0].text)
                    sheet1.cell(row=cell.row, column=2).value = re.sub(#品牌
                        r'\s', '', brand[0].text)
                    sheet1.cell(row=cell.row, column=3).value = re.sub(#型号
                        r'\s', '', mdel[0].text)

                    sheet1.cell(row=cell.row, column=11).value = re.sub(
                        r'\s|￥', '', quantity[0].text)

                    if re.sub(r'\s|￥', '', quantity[0].text) != '报价':
                        sheet1.cell(row=cell.row, column=12).value = float(re.sub(
                            r'\s|￥', '', quantity[0].text))*DesignatorNum
                    else:
                        sheet1.cell(row=cell.row, column=12).value = re.sub(
                            r'\s|￥', '', quantity[0].text)
                    
                    sheet1.cell(row=cell.row, column=10).value = DesignatorNum#保存器件位号数量
                    sheet1.cell(row=cell.row, column=8).value = re.sub(#封装
                        r'\s', '', package[0].text)
                    sheet1.cell(row=cell.row, column=13).value = '立创商城'#来源
                    #填充白色背景
                    for row in sheet1[cell.row]:
                        row.fill = PatternFill(
                            start_color="ffffff", fill_type="solid")
                except IndexError:
                    for row in sheet1[cell.row]:
                        row.fill = PatternFill(
                            start_color="92D050", fill_type="solid")
                # driver.quit()
                time.sleep(1)  # 延时一秒
            driver.quit()

if __name__ == '__main__':
    set_excel()
    get_data()
    # 保存表格
    if "_Copy" in excel_name:
        wb.save(filename=excel_name+str(".xlsx"))
    else:
        wb.save(filename=excel_name+str("_Copy.xlsx"))
    input("Press enter to exit...")

#打包成exe用： 
# pyinstaller -F --icon=lbxx.ico .\立创商城BOM表爬虫工具.py
# pyinstaller -F --add-binary="msedgedriver.exe;."  .\模拟用户访问浏览器.py把驱动一起打包进去
#还需解决不同浏览器版本之间不同驱动的问题
